import re
import sys
from pathlib import Path

sys.path.insert(0, "/private/tmp/lfl_pydeps")

from openpyxl import load_workbook, Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


SOURCE = Path("LFL Breakout Allocations_FINAL.xlsx")
OUTPUT = Path("LFL_Searchable_Session_Lookup.xlsx")


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def parse_session(text):
    text = clean(text)
    venue_match = re.search(r"Venue:\s*(.+)", text, re.I)
    venue = venue_match.group(1).strip() if venue_match else ""
    presenter_match = re.search(r"Presenters?:\s*(.+?)(?:\n|$)", text, re.I)
    presenter = presenter_match.group(1).strip() if presenter_match else ""
    title = re.split(r"\nPresenters?:|\n\s*Venue:", text, flags=re.I)[0].strip()
    title = re.sub(r"^\d+\.\s*", "", title)
    return title, presenter, venue


def session_assignments(ws):
    result = {}
    current = ""
    for row in range(2, ws.max_row + 1):
        name = clean(ws.cell(row, 2).value)
        details = clean(ws.cell(row, 3).value)
        if details:
            current = details
        if name:
            result[name] = parse_session(current)
    return result


source = load_workbook(SOURCE, data_only=False)
session_a = session_assignments(source["BREAKOUT SESSION A"])
session_b = session_assignments(source["BREAKOUT SESSION B"])

departments = []
dept_ws = source["DEPARTMENT REFLECTION"]
for row in range(2, dept_ws.max_row + 1):
    department = clean(dept_ws.cell(row, 1).value)
    if department:
        departments.append(
            (
                department,
                clean(dept_ws.cell(row, 2).value),
                clean(dept_ws.cell(row, 3).value),
            )
        )

names = sorted(set(session_a) | set(session_b), key=str.casefold)

wb = Workbook()
search = wb.active
search.title = "SEARCH"
directory = wb.create_sheet("PARTICIPANT DIRECTORY")
dept = wb.create_sheet("DEPARTMENT REFLECTION")
notes = wb.create_sheet("DATA NOTES")

navy = "17365D"
blue = "DCE6F1"
light_blue = "EAF2F8"
gold = "F4B183"
green = "E2F0D9"
grey = "E7E6E6"
white = "FFFFFF"
red = "FCE4D6"
thin = Side(style="thin", color="B7B7B7")

# Search dashboard
search.sheet_view.showGridLines = False
search.merge_cells("A1:H1")
search["A1"] = "LFL 2026 — Personal Session Finder"
search["A1"].font = Font(size=20, bold=True, color=white)
search["A1"].fill = PatternFill("solid", fgColor=navy)
search["A1"].alignment = Alignment(horizontal="center", vertical="center")
search.row_dimensions[1].height = 34

search.merge_cells("A2:H2")
search["A2"] = (
    "Type a full or partial name in B4. The first matching participant will appear below."
)
search["A2"].font = Font(italic=True, color="44546A")
search["A2"].alignment = Alignment(horizontal="center")

search["A4"] = "Search name"
search["A4"].font = Font(bold=True, color=white)
search["A4"].fill = PatternFill("solid", fgColor=navy)
search["B4"] = ""
search.merge_cells("B4:F4")
for row in search["B4:F4"]:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=gold)
        cell.border = Border(bottom=Side(style="medium", color=navy))
search["B4"].font = Font(size=14, bold=True)

search["A5"] = "Select department"
search["A5"].font = Font(bold=True, color=white)
search["A5"].fill = PatternFill("solid", fgColor=navy)
search["B5"] = ""
search.merge_cells("B5:F5")
for row in search["B5:F5"]:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=green)
        cell.border = Border(bottom=Side(style="medium", color=navy))
search["B5"].font = Font(size=12, bold=True)

search["A6"] = "Participant"
search["B6"] = '=IF($B$4="","",IFERROR(XLOOKUP("*"&$B$4&"*",\'PARTICIPANT DIRECTORY\'!$A$2:$A$200,\'PARTICIPANT DIRECTORY\'!$A$2:$A$200,"No matching name",2),"No matching name"))'
search.merge_cells("B6:H6")

headers = ["Session", "Time", "Session title", "Presenter(s)", "Venue", "Department / Notes", "", ""]
for col, value in enumerate(headers, 1):
    cell = search.cell(8, col, value)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.font = Font(bold=True, color=white)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

rows = [
    (
        "Breakout 1",
        "10:00–10:45",
        4,
        5,
        6,
        None,
    ),
    (
        "Breakout 2",
        "11:00–11:45",
        9,
        10,
        11,
        None,
    ),
    (
        "Department Reflection",
        "11:50–12:30",
        None,
        13,
        12,
        10,
    ),
]

for r, (label, time, title_col, presenter_col, venue_col, dept_col) in enumerate(rows, 9):
    search.cell(r, 1, label)
    search.cell(r, 2, time)
    if title_col:
        search.cell(
            r,
            3,
            f'=IFERROR(XLOOKUP($B$6,\'PARTICIPANT DIRECTORY\'!$A$2:$A$200,\'PARTICIPANT DIRECTORY\'!${chr(64+title_col)}$2:${chr(64+title_col)}$200,""),"")',
        )
        search.cell(
            r,
            4,
            f'=IFERROR(XLOOKUP($B$6,\'PARTICIPANT DIRECTORY\'!$A$2:$A$200,\'PARTICIPANT DIRECTORY\'!${chr(64+presenter_col)}$2:${chr(64+presenter_col)}$200,""),"")',
        )
        search.cell(
            r,
            5,
            f'=IFERROR(XLOOKUP($B$6,\'PARTICIPANT DIRECTORY\'!$A$2:$A$200,\'PARTICIPANT DIRECTORY\'!${chr(64+venue_col)}$2:${chr(64+venue_col)}$200,""),"")',
        )
        search.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    else:
        search.cell(r, 3, "Department Reflection & Consolidation")
        search.cell(
            r,
            4,
            '=IFERROR(XLOOKUP($B$5,\'DEPARTMENT REFLECTION\'!$A$2:$A$20,\'DEPARTMENT REFLECTION\'!$C$2:$C$20,""),"")',
        )
        search.cell(
            r,
            5,
            '=IFERROR(XLOOKUP($B$5,\'DEPARTMENT REFLECTION\'!$A$2:$A$20,\'DEPARTMENT REFLECTION\'!$B$2:$B$20,""),"")',
        )
        search.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        search.cell(
            r,
            6,
            '=$B$5',
        )
    for c in range(1, 9):
        cell = search.cell(r, c)
        cell.fill = PatternFill("solid", fgColor=light_blue if r % 2 else white)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    search.row_dimensions[r].height = 68

search.merge_cells("A13:H14")
search["A13"] = (
    "Choose your department from the green dropdown above to display the correct "
    "Department Reflection venue and facilitator."
)
search["A13"].fill = PatternFill("solid", fgColor=green)
search["A13"].font = Font(color="375623", italic=True)
search["A13"].alignment = Alignment(wrap_text=True, vertical="center")

for col, width in {"A": 22, "B": 17, "C": 48, "D": 28, "E": 15, "F": 18, "G": 14, "H": 14}.items():
    search.column_dimensions[col].width = width
search.freeze_panes = "A8"
search.auto_filter.ref = "A8:H11"

# Participant directory
directory_headers = [
    "Name",
    "Breakout 1 Time",
    "Breakout 1 Session",
    "Breakout 1 Title",
    "Breakout 1 Presenter(s)",
    "Breakout 1 Venue",
    "Breakout 2 Time",
    "Breakout 2 Session",
    "Breakout 2 Title",
    "Breakout 2 Presenter(s)",
    "Breakout 2 Venue",
    "Department",
    "Department Reflection Venue",
    "Department Reflection Facilitator",
]
for col, value in enumerate(directory_headers, 1):
    cell = directory.cell(1, col, value)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.font = Font(bold=True, color=white)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row, name in enumerate(names, 2):
    a_title, a_presenter, a_venue = session_a.get(name, ("", "", ""))
    b_title, b_presenter, b_venue = session_b.get(name, ("", "", ""))
    values = [
        name,
        "10:00–10:45",
        "Breakout 1",
        a_title,
        a_presenter,
        a_venue,
        "11:00–11:45",
        "Breakout 2",
        b_title,
        b_presenter,
        b_venue,
        "",
        f'=IFERROR(XLOOKUP(L{row},\'DEPARTMENT REFLECTION\'!$A$2:$A$20,\'DEPARTMENT REFLECTION\'!$B$2:$B$20,""),"")',
        f'=IFERROR(XLOOKUP(L{row},\'DEPARTMENT REFLECTION\'!$A$2:$A$20,\'DEPARTMENT REFLECTION\'!$C$2:$C$20,""),"")',
    ]
    for col, value in enumerate(values, 1):
        cell = directory.cell(row, col, value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=Side(style="hair", color="D9E1F2"))
    directory.cell(row, 12).fill = PatternFill("solid", fgColor=gold)

directory.freeze_panes = "A2"
directory.auto_filter.ref = f"A1:N{len(names)+1}"
table = Table(displayName="ParticipantDirectory", ref=f"A1:N{len(names)+1}")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
)
directory.add_table(table)
widths = [31, 16, 13, 52, 28, 15, 16, 13, 52, 28, 15, 48, 20, 30]
for col, width in enumerate(widths, 1):
    directory.column_dimensions[chr(64 + col)].width = width

# Department dropdown
validation = DataValidation(
    type="list",
    formula1="'DEPARTMENT REFLECTION'!$A$2:$A$15",
    allow_blank=True,
)
validation.promptTitle = "Choose department"
validation.prompt = "Choose the participant's department to populate reflection details."
validation.errorTitle = "Invalid department"
validation.error = "Please choose a department from the list."
directory.add_data_validation(validation)
validation.add(f"L2:L{len(names)+1}")
directory.conditional_formatting.add(
    f"L2:L{len(names)+1}",
    FormulaRule(formula=["L2=\"\""], fill=PatternFill("solid", fgColor=red)),
)

wb.defined_names.add(
    DefinedName(
        "DepartmentList",
        attr_text=f"'DEPARTMENT REFLECTION'!$A$2:$A${len(departments)+1}",
    )
)
search_dept_validation = DataValidation(
    type="list", formula1="=DepartmentList", allow_blank=False
)
search_dept_validation.promptTitle = "Select your department"
search_dept_validation.prompt = "Choose a department to show the reflection details."
search_dept_validation.errorTitle = "Invalid department"
search_dept_validation.error = "Please select a department from the dropdown list."
search.add_data_validation(search_dept_validation)
search_dept_validation.add(search["B5"])

# Department reference
dept_headers = ["Department", "Venue", "HOD / Facilitator"]
for col, value in enumerate(dept_headers, 1):
    cell = dept.cell(1, col, value)
    cell.fill = PatternFill("solid", fgColor=navy)
    cell.font = Font(bold=True, color=white)
for row, values in enumerate(departments, 2):
    for col, value in enumerate(values, 1):
        dept.cell(row, col, value)
        dept.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
dept_table = Table(displayName="DepartmentReflection", ref=f"A1:C{len(departments)+1}")
dept_table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
dept.add_table(dept_table)
dept.freeze_panes = "A2"
dept.column_dimensions["A"].width = 68
dept.column_dimensions["B"].width = 20
dept.column_dimensions["C"].width = 28

for ws in (directory, dept):
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 34

# Source-data checks
notes.sheet_view.showGridLines = False
notes.merge_cells("A1:D1")
notes["A1"] = "Source Data Notes"
notes["A1"].fill = PatternFill("solid", fgColor=navy)
notes["A1"].font = Font(size=16, bold=True, color=white)
notes["A1"].alignment = Alignment(horizontal="center")
notes.row_dimensions[1].height = 30
notes.append(
    [
        "Issue",
        "Participant",
        "Impact",
        "Recommended action",
    ]
)
for cell in notes[2]:
    cell.fill = PatternFill("solid", fgColor=grey)
    cell.font = Font(bold=True)
missing_a = sorted(set(session_b) - set(session_a), key=str.casefold)
missing_b = sorted(set(session_a) - set(session_b), key=str.casefold)
for name in missing_a:
    notes.append(
        [
            "Missing Breakout 1 allocation in source",
            name,
            "Breakout 1 details are blank in the search result.",
            "Confirm the allocation and update columns D–F in PARTICIPANT DIRECTORY.",
        ]
    )
for name in missing_b:
    notes.append(
        [
            "Missing Breakout 2 allocation in source",
            name,
            "Breakout 2 details are blank in the search result.",
            "Confirm the allocation and update columns I–K in PARTICIPANT DIRECTORY.",
        ]
    )
notes.append(
    [
        "Department selected by each user",
        "All participants",
        "Reflection details appear after the user selects a department.",
        "Use the green department dropdown on the SEARCH sheet.",
    ]
)
for row in notes.iter_rows(min_row=3):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)
notes.column_dimensions["A"].width = 42
notes.column_dimensions["B"].width = 32
notes.column_dimensions["C"].width = 48
notes.column_dimensions["D"].width = 66
notes.freeze_panes = "A3"

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(OUTPUT)
print(f"Created {OUTPUT} with {len(names)} participants and {len(departments)} departments.")
