import json
import re
import sys
from pathlib import Path

sys.path.insert(0, "/private/tmp/lfl_pydeps")

from openpyxl import load_workbook


SOURCE = Path("LFL Breakout Allocations_FINAL.xlsx")
OUTPUT = Path("participant-data.js")


def clean(value):
    return "" if value is None else str(value).replace("\xa0", " ").strip()


def parse_details(value):
    text = clean(value)
    venue_match = re.search(r"Venue:\s*(.+)", text, re.I)
    presenter_match = re.search(r"Presenters?:\s*(.+?)(?:\n|$)", text, re.I)
    title = re.split(r"\nPresenters?:|\n\s*Venue:", text, flags=re.I)[0].strip()
    return {
        "title": re.sub(r"^\d+\.\s*", "", title),
        "presenter": presenter_match.group(1).strip() if presenter_match else "",
        "venue": venue_match.group(1).strip() if venue_match else "",
    }


def assignments(sheet):
    result = {}
    current = ""
    for row in range(2, sheet.max_row + 1):
        name = clean(sheet.cell(row, 2).value)
        details = clean(sheet.cell(row, 3).value)
        if details:
            current = details
        if name:
            result[name] = parse_details(current)
    return result


workbook = load_workbook(SOURCE, data_only=False, read_only=True)
first = assignments(workbook["BREAKOUT SESSION A"])
second = assignments(workbook["BREAKOUT SESSION B"])

participants = []
for name in sorted(set(first) | set(second), key=str.casefold):
    participants.append(
        {
            "name": name,
            "session1": first.get(name),
            "session2": second.get(name),
        }
    )

departments = []
sheet = workbook["DEPARTMENT REFLECTION"]
for row in range(2, sheet.max_row + 1):
    name = clean(sheet.cell(row, 1).value)
    if name:
        departments.append(
            {
                "name": name,
                "venue": clean(sheet.cell(row, 2).value),
                "facilitator": clean(sheet.cell(row, 3).value),
            }
        )

payload = {"participants": participants, "departments": departments}
OUTPUT.write_text(
    "window.LFL_DATA = "
    + json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    + ";\n",
    encoding="utf-8",
)
print(f"Exported {len(participants)} participants and {len(departments)} departments.")
